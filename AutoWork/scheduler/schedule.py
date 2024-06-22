import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import openpyxl as op
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

dfs = []

# 엑셀 파일 경로
path = 'schedules.xlsx'

date_range = {
    '2023' : [str(i) for i in range(11, 13)],
    '2024' : [str(i) for i in range(1, 13)],
    '2025' : [str(i) for i in range(1, 3)],
}

urls = [
    'https://www.seoultech.ac.kr/life/sch/common/',
    'https://www.seoultech.ac.kr/life/sch/grad/',
    'https://en.seoultech.ac.kr/acad/under/schedule/',
    'https://en.seoultech.ac.kr/acad/graduate/schedule/',
    'https://cn.seoultech.ac.kr/acad/under/schedule/',
    'https://cn.seoultech.ac.kr/acad/graduate/schedule/',
]

for i in range(6):
    data = []
    for year in date_range:
        for month in date_range[year]:
            print(f"@({year}/{month}) 데이터 추출")
            # 웹사이트 URL
            querystring = f'?mon={month}&year={year}'
            url = urls[i] + querystring
            # 요청 및 응답 받기
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')

            # 일정표 데이터 추출 (예: 테이블 형태의 데이터)
            table = soup.find('table', {'class': 'schedule'})
            rows = table.find_all('tr')

            date = soup.find('span', {'class': 'day'})
            date = date.text.strip()

            # 데이터 정제
            check = 0
            for row in rows[1:]:
                if row:
                    if check:
                        cols = row.find_all('td')
                        cols = [ele.text.strip() for ele in cols]
                        data.append([''] + [ele for ele in cols if ele])
                    else:
                        cols = row.find_all('td')
                        cols = [ele.text.strip() for ele in cols]
                        data.append([date] + [ele for ele in cols if ele])
                        check = 1
            # data.append(['', '', ''])

    # DataFrame 생성
    df = pd.DataFrame(data, columns=['date', 'date_range', 'text'])
    dfs.append(df)

print("@데이터 저장 중")

sheetnames = [
    '대학일정',
    '대학원일정',
    '대학일정(영문)',
    '대학원일정(영문)',
    '대학일정(중문)',
    '대학원일정(중문)',
]

# ExcelWriter 객체 생성
with pd.ExcelWriter(path, engine='openpyxl') as writer:
    for i in range(6):
        dfs[i].to_excel(writer, sheet_name=sheetnames[i], index=False)

wb = op.load_workbook(path)

cols_width = [10, 30, 80]

for ws in wb:  # 모든 시트 순회
    for i in range(3):  # 첫 세 열에 대해서만 반복
        column_letter = get_column_letter(i + 1)  # 열 번호를 열 문자로 변환
        ws.column_dimensions[column_letter].width = cols_width[i]  # 열 너비 설정

    for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
wb.save(path)
print("@데이터 저장 완료")
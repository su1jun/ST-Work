from google_play_scraper import Sort, reviews_all
from app_store_scraper import AppStore
import pandas as pd
import xmltodict
import requests
from datetime import datetime

import openpyxl as op
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
from tkcalendar import DateEntry
from tkinter import * # __all__
from tkinter import filedialog
import os
import json

def get_google_reviews_all(app_name, lang, country):
    result = reviews_all(
        app_name,
        sleep_milliseconds=2000,  # defaults to 0
        lang=lang,  # defaults to 'en'
        country=country,  # defaults to 'us'
        sort=Sort.MOST_RELEVANT,  # defaults to Sort.MOST_RELEVANT
        filter_score_with=None  # defaults to None(means all score)
    )

    return pd.DataFrame(result)

def get_last_page_num(url):
    response = requests.get(url).content.decode('utf8')
    xml = xmltodict.parse(response)
    last_url = [l['@href'] for l in xml['feed']['link'] if (l['@rel'] == 'last')][0]
    last_index = [int(s.replace('page=', '')) for s in last_url.split('/') if ('page=' in s)][0]
    return last_index

def extract_reply_content(df_cloumn):
    if str(df_cloumn) != 'nan':
        res = df_cloumn['body']
    else:
        res = None
    return res

def extract_reply_at(df_cloumn):
    if str(df_cloumn) != 'nan':
        res = df_cloumn['modified']
    else:
        res = None
    return res

# 리뷰 답변 버전 사용 가능, 단 해당 버전은 사용시 설치시 버전을 확인 불가
def get_ios_reviews_all_add_responce(id):
    my_app = AppStore(country='kr', app_name='gccare', app_id=id)
    my_app.review()
    fetched_reviews = my_app.reviews
    ios_review_df = pd.DataFrame(fetched_reviews)
    ios_review_df['reviewId'] = 'ios_' + ios_review_df.reset_index()['index'].astype('str')
    ios_review_df.rename(columns={'date': 'at', 'rating': 'score', 'review': 'content'}, inplace=True)

    ios_review_df['replyContent'] = ios_review_df['developerResponse'].apply(extract_reply_content)
    ios_review_df['repliedAt'] = ios_review_df['developerResponse'].apply(extract_reply_at)
    ios_review_df = ios_review_df.drop(['isEdited', 'developerResponse'], axis=1)

    return ios_review_df

def get_ios_reviews_all(id, country):
    # rss link 양식
    # https://itunes.apple.com/국가명/rss/customerreviews/page=1/id=아이디명/sortBy=mostRecent/xml
    url = 'https://itunes.apple.com/' + country + '/rss/customerreviews/page=1/id=' + id + '/sortBy=mostRecent/xml'

    try:
        last_index = get_last_page_num(url)
    except Exception as e:
        # print(url)
        print('No Reviews: appid %i' % id)
        print('Exception:', e)
        return

    result = []
    # 각 review page별로 추가
    for idx in range(1, last_index + 1):
        url = 'https://itunes.apple.com/' + country + '/rss/customerreviews/page=' + str(
            idx) + '/id=' + id + '/sortBy=mostRecent/xml'
        response = requests.get(url).content.decode('utf8')
        xml = xmltodict.parse(response)

        try:
            review_content = xml['feed']['entry']
        except KeyError as e:
            print("Empty Review", e)
            continue

        # 단일 리뷰인 경우 xml 구조가 달라짐
        try:
            print(review_content[0]['author']['name'])
            single_review = False
        except:
            single_review = True
            pass

        if single_review:
            result.append({
                'reviewId': review_content['id'],
                'userName': review_content['author']['name'],
                'at': review_content['updated'],
                'score': int(review_content['im:rating']),
                'title': review_content['title'],
                'content': review_content['content'][0]['#text'],
                'reviewCreatedVersion': review_content['im:version'],
            })
        else:
            for i in range(len(review_content)):
                result.append({
                    'reviewId': review_content[i]['id'],
                    'userName': review_content[i]['author']['name'],
                    'at': review_content[i]['updated'],
                    'score': int(review_content[i]['im:rating']),
                    'title': review_content[i]['title'],
                    'content': review_content[i]['content'][0]['#text'],
                    'reviewCreatedVersion': review_content[i]['im:version'],
                })
    res_df = pd.DataFrame(result)
    res_df['at'] = pd.to_datetime(res_df['at'], format="%Y-%m-%dT%H:%M:%S-07:00")

    return res_df

def get_ios_reviews(ios_app_id, country):
    fir_df = get_ios_reviews_all(ios_app_id, country)
    sec_df = get_ios_reviews_all_add_responce(ios_app_id)

    merge_df = pd.merge(left=fir_df, right=sec_df, how="left", on="userName")
    merge_df = merge_df.drop(['score_y', 'at_y', 'title_y', 'content_y', 'reviewId_y'], axis=1)
    merge_df.rename(columns={'score_x': 'score', 'at_x': 'at', 'title_x': 'title', 'content_x': 'content',
                             'reviewId_x': 'reviewId'}, inplace=True)
    merge_df['repliedAt'] = pd.to_datetime(merge_df['repliedAt'], format="%Y-%m-%dT%H:%M:%SZ")

    return merge_df

##### 엑셀 관련 함수
# 행열 정보 가져오기
def number_to_coordinate(rc):
    row_idx, col_idx = rc[0], rc[1]
    col_string = get_column_letter(col_idx)
    return f'{col_string}{row_idx}'

# 조건부 서식
def conditionFormat(ws, max_col):
    max_row  = ws.max_row
    for row_index in range(2, max_row+1):
        if row_index % 2 == 0:
            for col_index in range(2, max_col+1):
                ws.cell(row=row_index, column=col_index).fill = PatternFill(fill_type='solid', start_color='eeeeee', end_color='eeeeee')

# 저장 경로 (폴더)
def browse_dest_path():
    folder_selected = filedialog.askdirectory()
    if folder_selected == "": # 사용자가 취소를 누를 때
        print("폴더 선택 취소")
        return
    #print(folder_selected)
    txt_dest_path.delete(0, END)
    txt_dest_path.insert(0, folder_selected)

# 시작
def start():
    country = 'kr'
    lan = 'ko'
    
    ios_app_id = frame1_ety1.get()
    google_app_name = frame1_ety2.get()

    start_date = frame2_det1.get_date()
    end_date = frame2_det2.get_date()

    file_path = txt_dest_path.get()
    if file_path == "": path = txt_name_path.get() + Todate.get() + ".xlsx"
    else: path = txt_dest_path.get() + '/' + txt_name_path.get() + Todate.get() + ".xlsx"


    # ios 리뷰 df 추출
    ios_review_df = get_ios_reviews(ios_app_id, country)
    
    ios_review_df['os'] = 'ios'
    
    # 구글 리뷰 df 추출
    google_review_df = get_google_reviews_all(google_app_name, lan, country)
    google_review_df['os'] = 'google'

    # ios google 통합
    all_review_df = pd.concat([ios_review_df, google_review_df])

    try:
        start_date = frame2_det1.get_date()
        end_date = frame2_det2.get_date()
    except ValueError:
        msgbox.showinfo('정보', '날짜 입력이 잘못됐습니다.')

    if start_date > end_date:
        msgbox.showinfo('정보', '시작일이 더 큽니다.')

    # 날짜 포멧팅
    all_review_df['at'] = all_review_df['at'].dt.date
    all_review_df['repliedAt'] = all_review_df['repliedAt'].dt.date
    all_review_df = all_review_df[all_review_df['at'].between(start_date, end_date)]

    # 열 이름 바꾸기
    all_review_df.columns = ['ID', '작성자', '작성일시', '별점', '제목', '리뷰내용', '사용버전', '답변', '답변일시', 'OS', '작성자 이미지(google)', '좋아요(google)']

    # ID Drop
    all_review_df['ID'] = [' ' for _ in range(all_review_df.shape[0])]


    ios_review_df = all_review_df[all_review_df.OS == 'ios']
    ios_review_df = ios_review_df.sort_values('작성일시', ascending=False)
    ios_review_df = ios_review_df.reset_index(drop=True)
    ios_review_df.index += 1

    google_review_df = all_review_df[all_review_df.OS == 'google']
    google_review_df = google_review_df.sort_values('작성일시', ascending=False)
    google_review_df = google_review_df.reset_index(drop=True)
    google_review_df.index += 1

    # 데이터 저장
    with pd.ExcelWriter(path) as writer:
        ios_review_df.to_excel(writer, sheet_name='App Store')
        google_review_df.to_excel(writer, sheet_name='Google play Store')

    wb = op.load_workbook(path)
    
    # 셀 너비, 높이 설정하기
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    cols_dim1 = [4, 5, 21, 12, 5, 20, 85, 10, 50, 12, 8, 25, 15]
    cols_dim2 = [4, 5, 21, 12, 5, 5, 100, 10, 50, 12, 8, 25, 15]

    ws = wb['App Store']
    for i, col in enumerate(cols):
        ws.column_dimensions[col].width = cols_dim1[i]
        for i in range(2, ios_review_df.shape[0]+2):
            ws[col+str(i)].alignment = Alignment(horizontal = 'center', vertical='center', wrap_text=True)
    
    ws = wb['Google play Store']
    for i, col in enumerate(cols):
        ws.column_dimensions[col].width = cols_dim2[i]
        for i in range(2, google_review_df.shape[0]+2):
            ws[col+str(i)].alignment = Alignment(horizontal = 'center', vertical='center',  wrap_text=True)

    for os in ['App Store', 'Google play Store']:
        ws = wb[os]
        ## 필터 범위
        start_row, start_col = 1, 1
        end_row, end_col = all_review_df.shape[0], 13
        
        cell_range = f'{number_to_coordinate((start_row, start_col))}:{number_to_coordinate((end_row, end_col))}'
        ws.auto_filter.ref = cell_range ## 필터 범위 지정
        ws.auto_filter.add_filter_column(0, []) ## 필터 생성

        for col_index in range(2, 14):
            ws.cell(row=1, column=col_index).fill = PatternFill(fill_type='solid', start_color='cccccc', end_color='cccccc')
        max_row  = ws.max_row
        for row_index in range(1, max_row+1):
            ws.cell(row=row_index, column=1).fill = PatternFill(fill_type='solid', start_color='cccccc', end_color='cccccc')
        conditionFormat(ws, 13)

    #Save
    # print(all_review_df.dtypes)
    wb.save(path)
    wb.close()

    msgbox.showinfo('정보', '작업이 끝났습니다.')

def on_exit():
    """
    Event handler for when the program exits.
    Saves the value of the date_entry_widget to a file.
    """
    data = dict()
    data['ios_id'] = frame1_ety1.get()
    data['google_id'] = frame1_ety2.get()
    data['start_date'] = frame2_det1.get_date()
    data['name'] = txt_name_path.get()
    data['path'] = txt_dest_path.get()
    save_last_date(data)
    #print("data", data)
    root.destroy()

if __name__ == "__main__":
    root = Tk()
    root.title("App reivew Crawler")

    # 옵션 프레임
    action_frame = Frame(root)
    action_frame.pack(fill="x", padx=5, pady=5) # 간격 띄우기

    # 앱 ID 입력
    frame1 = LabelFrame(action_frame, text="APP ID")
    frame1.pack(fill="x", padx=5, pady=5, ipady=8)

    frame1_lbl1 = Label(frame1, text="iPhone :")
    frame1_lbl1.pack(side="left", padx=6, pady=2)

    frame1_ety1 = Entry(frame1)
    frame1_ety1.pack(side="left", fill="x", expand=True, padx=1, pady=2) # 높이 변경

    frame1_lbl3 = Label(frame1, text="Android :")
    frame1_lbl3.pack(side="left", padx=6, pady=2)

    frame1_ety2 = Entry(frame1)
    frame1_ety2.pack(side="left", fill="x", expand=True, padx=1, pady=2) # 높이 변경
    
    # 다운로드 날짜
    frame2 = LabelFrame(action_frame, text="검색 기간")
    frame2.pack(fill="x", padx=5, pady=5, ipady=8)

    frame2_lbl1 = Label(frame2, text="Start Date :")
    frame2_lbl1.pack(side="left", padx=6, pady=2)

    frame2_det1 = DateEntry(frame2, state='normal', date_pattern='yyyy/mm/dd')
    frame2_det1.pack(side="left", fill="x", expand=True, padx=1, pady=2)

    frame2_lbl3 = Label(frame2, text="End Date :")
    frame2_lbl3.pack(side="left", padx=6, pady=2)

    frame2_det2 = DateEntry(frame2, state='normal', date_pattern='yyyy/mm/dd')
    frame2_det2.pack(side="left", fill="x", expand=True, padx=1, pady=2)

    # 파일 이름
    frame3 = LabelFrame(action_frame, text="파일 이름")
    frame3.pack(fill="x", padx=5, pady=5, ipady=8)

    frame3_lbl1 = Label(frame3, text="제목 :")
    frame3_lbl1.pack(side="left", padx=6, pady=2)

    txt_name_path = Entry(frame3)
    txt_name_path.pack(side="left", fill="x", expand=True, padx=8, pady=5, ipady=4)

    frame3_lbl2 = Label(frame3, text="+ :")
    frame3_lbl2.pack(side="left", padx=20, pady=2)

    Todate = Entry(frame3)
    Todate.pack(side="left", fill="x", expand=True, padx=2, pady=5, ipady=4)
    Todate.insert(0, datetime.today().strftime('_%y%m%d'))

    # 저장 경로 프레임
    path_frame = LabelFrame(root, text="저장 경로")
    path_frame.pack(fill="x", padx=5, pady=5, ipady=5)

    txt_dest_path = Entry(path_frame)
    txt_dest_path.pack(side="left", fill="x", expand=True, padx=5, pady=5, ipady=4) # 높이 변경

    btn_dest_path = Button(path_frame, text="찾아보기", width=10, command=browse_dest_path)
    btn_dest_path.pack(side="right", padx=5, pady=5)

    # 실행 프레임
    frame_run = Frame(root)
    frame_run.pack(fill="x", padx=5, pady=5)

    btn_close = Button(frame_run, padx=5, pady=5, text="닫기", width=12, command=on_exit)
    btn_close.pack(side="right", padx=5, pady=5)

    btn_start = Button(frame_run, padx=5, pady=5, text="시작", width=12, command=start)
    btn_start.pack(side="right", padx=5, pady=5)

    root.resizable(False, False)

    SAVE_FILE_PATH = '_default_.txt'

    def save_last_date(pre_data):
        """
        Saves the specified date string to a file.
        """
        pre_data['start_date'] = pre_data['start_date'].strftime('%Y/%m/%d')

        if os.path.exists(SAVE_FILE_PATH):
            with open(SAVE_FILE_PATH, 'w') as f:
                f.write(json.dumps(pre_data))
        else:
            with open(SAVE_FILE_PATH, mode="w+") as f:
                f.write(json.dumps(pre_data))
        print(f"File path: {os.path.abspath(SAVE_FILE_PATH)}")

    def load_last_date():
        """
        Loads the last saved date string from the file.
        Returns None if the file does not exist or is empty.
        """

        if os.path.exists(SAVE_FILE_PATH):
            with open(SAVE_FILE_PATH, 'r') as f:
                last_data = json.loads(f.read())
                if last_data:
                    return last_data
        return {'ios_id' : '966808150', 'google_id' : 'ac.kr.seoultech.m2', 'start_date' : '2023/02/13', 'name' : 'result', 'path' : ''}
    frame1_ety1.insert('0', '966808150')
    frame1_ety2.insert('0', 'ac.kr.seoultech.m2')

    # load the last saved date value from the file
    last_date = load_last_date()
    if last_date['ios_id'] != '':
        frame1_ety1.delete(0,"end")
        frame1_ety1.insert('0', last_date['ios_id'])
    if last_date['google_id'] != '':
        frame1_ety2.delete(0,"end")
        frame1_ety2.insert('0', last_date['google_id'])
    if last_date['start_date'] is not None: frame2_det1.set_date(datetime.strptime(last_date['start_date'], '%Y/%m/%d'))
    if last_date['name'] is not None: txt_name_path.insert('0', last_date['name'])
    if last_date['path'] is not None: txt_dest_path.insert('0', last_date['path'])

    root.protocol("WM_DELETE_WINDOW", on_exit)
    root.mainloop()

    # root.protocol("WM_DELETE_WINDOW", on_closing)

# ios_review_df.columns = ['ID', '작성자', '작성일시', '별점', '제목', '리뷰내용', '사용버전', 'OS']
# ios_review_df.to_csv("gccare_ios_review_0705.csv", encoding='utf-8-sig', index=False)
# google_review_df.columns = ['ID', '작성자', '작성일시', '별점', '제목', '리뷰내용', '사용버전', '작성자 이미지(google)', '좋아요(google)', '답변(google)', '답변일시(google)', 'OS']
# google_review_df.to_csv("gccare_google_review_0705.csv", encoding='utf-8-sig', index=False)